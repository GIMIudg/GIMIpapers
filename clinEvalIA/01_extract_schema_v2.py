"""
PASO 1: Extracción del schema del formulario

Tiene dos modos controlados por la variable MODE:

  MODE = "evaluation"  — extrae TODAS las preguntas select normales
                         (fuera Y dentro de repeats) para el LLM-as-judge.
                         Produce: schema_evaluation.json (306 preguntas)

  MODE = "generation"  — extrae solo preguntas fuera de repeats + gate questions
                         para la generación sintética tabular.
                         Produce: schema_generation.json (176 preguntas)

Estrategia para from_file (ambos modos):
  - Estados INEGI     → incluidos (32 entidades)
  - Municipios INEGI  → excluidos (~2500 valores)
  - SINCO ocupación   → reducido a 9 grupos principales
  - ICD-10 / ICD-11   → excluidos
  - RxNorm / LOINC    → excluidos
  - ISO países        → reducido a 15 países frecuentes en MX
"""

import json
import openpyxl
from collections import defaultdict

# ── Configuración ──────────────────────────────────────────────────────────────
EXCEL_PATH = "ruta_al_formulario_excel"
MODE       = "evaluation"   # "evaluation" o "generation"
OUTPUT_PATH = f"schema_{MODE}_ver_mejorada.json"

# ── Secciones para modo generation ────────────────────────────────────────────
GENERATION_SECTIONS = [
    'ROOT',
    'SECCIÓN 8.1. Cirugías',
    'SECCIÓN 8.2. Alergias',
    'SECCIÓN 8.3. Transfusiones',
    'SECCIÓN 9.1. Enfermedades cardiovasculares',
    'SECCIÓN 9.2. Enfermedades cerebrovasculares',
    'SECCIÓN 9.3. Diabetes mellitus',
    'SECCIÓN 9.4. Hipertensión arterial',
    'SECCIÓN 9.5. Dislipidemias',
    'SECCIÓN 9.6. Enfermedades pulmonares',
    'SECCIÓN 9.7. Enfermedades renales',
    'SECCIÓN 9.8. Enfermedades hepáticas',
    '9.8.1.a CIRROSIS HEPÁTICA',
    '9.8.1.b HEPATITIS CRÓNICA',
    '9.8.1.c HÍGADO GRASO',
    'SECCIÓN 9.9. Enfermedades del tejido conectivo y autoinmunes',
    'SECCIÓN 9.10. Úlcera péptica',
    'SECCIÓN 9.11. Demencia',
    'SECCIÓN 9.12. Cáncer (tumores sólidos)',
    'SECCIÓN 9.13. Leucemia y linfomas',
    'Leucemia',
    'Linfoma/Mieloma',
    'SECCIÓN 9.14. SIDA/VIH',
    'SECCIÓN 9.15. Otras enfermedades crónicas',
    '10. Consulta actual y síntomas',
    '11. Medicamentos actuales',
    '13. Estudios complementarios',
    'Resultado de valoración/índice (opcional)',
]

# Gate questions para modo generation
GATE_QUESTIONS = {
    'agregar_fam':             'Registro de familiares con enfermedad',
    'hay_sustancias':          'Registro de sustancias',
    'convive_animales':        'Registro de animales',
    'expo_tuvo':               'Registro de exposiciones',
    'vac_esquema_completo':    'Registro de vacunas',
    'ss2_uso_anticonceptivos': 'Registro de métodos anticonceptivos',
    'ap9_tiene_cirugias':      'Registro de cirugías',
    'ap9_tiene_alergias':      'Registro de alergias',
    'ap9_tiene_transfusiones': 'Registro de transfusiones',
    'ap10_tipo_cancer':        'Detalle por tipo de cáncer',
    'ap10_otras_enf_list':     'Otras enfermedades crónicas',
    's12_toma_medicamentos':   'Registro de medicamentos',
    's13_hay_estudios_comp':   'Registro de estudios complementarios',
}

# ── Catálogos reducidos ────────────────────────────────────────────────────────
ESTADOS_INEGI = [
    {"value": "01", "label": "Aguascalientes"},
    {"value": "02", "label": "Baja California"},
    {"value": "03", "label": "Baja California Sur"},
    {"value": "04", "label": "Campeche"},
    {"value": "05", "label": "Coahuila"},
    {"value": "06", "label": "Colima"},
    {"value": "07", "label": "Chiapas"},
    {"value": "08", "label": "Chihuahua"},
    {"value": "09", "label": "Ciudad de México"},
    {"value": "10", "label": "Durango"},
    {"value": "11", "label": "Guanajuato"},
    {"value": "12", "label": "Guerrero"},
    {"value": "13", "label": "Hidalgo"},
    {"value": "14", "label": "Jalisco"},
    {"value": "15", "label": "Estado de México"},
    {"value": "16", "label": "Michoacán"},
    {"value": "17", "label": "Morelos"},
    {"value": "18", "label": "Nayarit"},
    {"value": "19", "label": "Nuevo León"},
    {"value": "20", "label": "Oaxaca"},
    {"value": "21", "label": "Puebla"},
    {"value": "22", "label": "Querétaro"},
    {"value": "23", "label": "Quintana Roo"},
    {"value": "24", "label": "San Luis Potosí"},
    {"value": "25", "label": "Sinaloa"},
    {"value": "26", "label": "Sonora"},
    {"value": "27", "label": "Tabasco"},
    {"value": "28", "label": "Tamaulipas"},
    {"value": "29", "label": "Tlaxcala"},
    {"value": "30", "label": "Veracruz"},
    {"value": "31", "label": "Yucatán"},
    {"value": "32", "label": "Zacatecas"},
]

SINCO_GRUPOS = [
    {"value": "1", "label": "Funcionarios, directores y jefes"},
    {"value": "2", "label": "Profesionistas y técnicos"},
    {"value": "3", "label": "Trabajadores auxiliares en actividades administrativas"},
    {"value": "4", "label": "Comerciantes, empleados en ventas y agentes de ventas"},
    {"value": "5", "label": "Trabajadores en servicios personales y vigilancia"},
    {"value": "6", "label": "Trabajadores en actividades agrícolas, ganaderas y forestales"},
    {"value": "7", "label": "Trabajadores artesanales"},
    {"value": "8", "label": "Operadores de maquinaria industrial y transporte"},
    {"value": "9", "label": "Trabajadores en actividades elementales y de apoyo"},
]

PAISES_FRECUENTES = [
    {"value": "MX", "label": "México"},
    {"value": "US", "label": "Estados Unidos"},
    {"value": "GT", "label": "Guatemala"},
    {"value": "HN", "label": "Honduras"},
    {"value": "SV", "label": "El Salvador"},
    {"value": "CO", "label": "Colombia"},
    {"value": "VE", "label": "Venezuela"},
    {"value": "CU", "label": "Cuba"},
    {"value": "AR", "label": "Argentina"},
    {"value": "ES", "label": "España"},
    {"value": "CA", "label": "Canadá"},
    {"value": "DE", "label": "Alemania"},
    {"value": "FR", "label": "Francia"},
    {"value": "GB", "label": "Reino Unido"},
    {"value": "otro", "label": "Otro país"},
]

FROM_FILE_STRATEGY = {
    "ageeml_estados_inegi.csv":              ("include", ESTADOS_INEGI,     "Estados INEGI — 32 entidades"),
    "ageeml_municipios_inegi.csv":           ("exclude", [],                "Excluido: ~2500 municipios"),
    "sinco2019_grupos_unitarios.csv":        ("reduce",  SINCO_GRUPOS,      "SINCO reducido a 9 grupos"),
    "icd11_mms_es_en_kobo_leaf_v2.csv":      ("exclude", [],                "Excluido: ICD-11"),
    "catalogo_icd10_kobo.csv":               ("exclude", [],                "Excluido: ICD-10"),
    "iso3166_paises.csv":                    ("reduce",  PAISES_FRECUENTES, "ISO 3166 reducido a 15 países"),
    "rxnorm_prescribable_kobo_ingredientes_es_v1.csv": ("exclude", [], "Excluido: RxNorm"),
    "catalogo_efectos_adversos_s12_v1.csv":  ("exclude", [],                "Excluido: efectos adversos"),
    "loinc_img.csv":                         ("exclude", [],                "Excluido: LOINC imagen"),
    "lab_loinc.csv":                         ("exclude", [],                "Excluido: LOINC laboratorio"),
    "bio_loinc.csv":                         ("exclude", [],                "Excluido: LOINC bioseñales"),
    "val_indices.csv":                       ("exclude", [],                "Excluido: valoraciones"),
}

# ── Carga del Excel ────────────────────────────────────────────────────────────
wb         = openpyxl.load_workbook(EXCEL_PATH)
ws_survey  = wb['survey']
ws_choices = wb['choices']

choices_map = defaultdict(list)
for row in ws_choices.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1] and row[2]:
        choices_map[row[0]].append({'value': row[1], 'label': row[2]})

# ── Extracción ─────────────────────────────────────────────────────────────────
rows = list(ws_survey.iter_rows(min_row=2, values_only=True))

current_section = 'ROOT'
repeat_depth    = 0
repeat_name     = None
questions       = []
excluded_log    = []

for row in rows:
    qtype, qname, qlabel = row[0], row[1], row[2]
    relevant = row[8]

    if qtype == 'begin_group':
        current_section = qlabel or qname
    elif qtype == 'begin_repeat':
        repeat_depth += 1
        repeat_name = qlabel or qname
    elif qtype == 'end_repeat':
        repeat_depth = max(0, repeat_depth - 1)  # tolerante a end_repeat huérfano
        if repeat_depth == 0:
            repeat_name = None

    if not qtype or 'select' not in str(qtype):
        continue

    is_gate      = qname in GATE_QUESTIONS
    is_from_file = 'from_file' in str(qtype)
    is_multiple  = 'select_multiple' in str(qtype)
    in_repeat    = repeat_depth > 0
    in_target    = current_section in GENERATION_SECTIONS

    # ── from_file: aplicar estrategia en ambos modos ──────────────────────
    if is_from_file:
        file_name = str(qtype).split('from_file ')[-1].strip()
        strategy, reduced_choices, note = FROM_FILE_STRATEGY.get(
            file_name, ("exclude", [], f"Excluido: catálogo {file_name}")
        )
        if strategy == "exclude":
            excluded_log.append({'name': qname, 'label': qlabel, 'reason': note})
            continue
        questions.append({
            'section':     current_section,
            'name':        qname,
            'label':       qlabel,
            'type':        'select_multiple' if is_multiple else 'select_one',
            'choices':     reduced_choices,
            'relevant':    relevant,
            'required':    row[5] == 'yes',
            'category':    'from_file_reduced',
            'in_repeat':   in_repeat,
            'repeat_name': repeat_name,
            'note':        note
        })
        continue

    # ── MODO EVALUATION: incluir todo lo que no sea from_file ─────────────
    if MODE == "evaluation":
        list_name = (str(qtype)
                     .replace('select_one ', '')
                     .replace('select_multiple ', '')
                     .strip())
        questions.append({
            'section':     current_section,
            'name':        qname,
            'label':       qlabel,
            'type':        'select_multiple' if is_multiple else 'select_one',
            'choices':     choices_map.get(list_name, []),
            'relevant':    relevant,
            'required':    row[5] == 'yes',
            'category':    'select',
            'in_repeat':   in_repeat,
            'repeat_name': repeat_name,
            'note':        None
        })
        continue

    # ── MODO GENERATION ───────────────────────────────────────────────────

    # Gate question: incluir siempre fuera de repeat
    if is_gate and not in_repeat:
        list_name = (str(qtype)
                     .replace('select_one ', '')
                     .replace('select_multiple ', '')
                     .strip())
        questions.append({
            'section':     current_section,
            'name':        qname,
            'label':       qlabel,
            'type':        'select_multiple' if is_multiple else 'select_one',
            'choices':     choices_map.get(list_name, []),
            'relevant':    relevant,
            'required':    row[5] == 'yes',
            'category':    'gate',
            'in_repeat':   False,
            'repeat_name': GATE_QUESTIONS[qname],
            'note':        'Pregunta de entrada de sección iterativa'
        })
        continue

    # Dentro de repeat → excluir en modo generation
    if in_repeat:
        continue

    # Fuera de sección objetivo → excluir
    if not in_target:
        continue

    list_name = (str(qtype)
                 .replace('select_one ', '')
                 .replace('select_multiple ', '')
                 .strip())
    questions.append({
        'section':     current_section,
        'name':        qname,
        'label':       qlabel,
        'type':        'select_multiple' if is_multiple else 'select_one',
        'choices':     choices_map.get(list_name, []),
        'relevant':    relevant,
        'required':    row[5] == 'yes',
        'category':    'select',
        'in_repeat':   False,
        'repeat_name': None,
        'note':        None
    })

# ── Guardar ────────────────────────────────────────────────────────────────────
schema = {
    'form_title':      'HC — Historia Clínica',
    'mode':            MODE,
    'n_questions':     len(questions),
    'questions':       questions,
    'excluded_items':  excluded_log,
    'design_decisions': {
        'from_file_excluded': 'ICD-10, ICD-11, RxNorm, LOINC, municipios — catálogos con miles de entradas, inviables para evaluación LLM',
        'repeats_evaluation': 'Incluidas en modo evaluation para validar el instrumento completo',
        'repeats_generation': 'Solo gate question en modo generation — formato tabular no soporta estructura jerárquica',
    }
}

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(schema, f, ensure_ascii=False, indent=2)

# ── Reporte ────────────────────────────────────────────────────────────────────
from collections import Counter
cats = Counter(q['category'] for q in questions)
in_rep = sum(1 for q in questions if q['in_repeat'])

print(f"Modo:    {MODE}")
print(f"Output:  {OUTPUT_PATH}")
print(f"\nTotal incluidas:  {len(questions)}")
print(f"Total excluidas:  {len(excluded_log)}")
print(f"\nDentro de repeat: {in_rep}")
print(f"Fuera de repeat:  {len(questions) - in_rep}")
print(f"\nPor categoría:")
for cat, n in cats.most_common():
    print(f"  {cat}: {n}")

if MODE == "generation":
    print(f"\nSecciones incluidas ({len(GENERATION_SECTIONS)}):")
    sec_counts = Counter(q['section'] for q in questions if q['category'] == 'select')
    for sec in GENERATION_SECTIONS:
        n = sec_counts.get(sec, 0)
        if n > 0:
            print(f"  [{n:3d}] {sec}")
